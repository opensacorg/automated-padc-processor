import pandas as pd
import openpyxl
from tqdm import tqdm
import time


# Define the functions


def find_row_with_value(df, target_value):
    """
    Finds row numbers where the target value is located in the DataFrame.

    Args:
    - df (DataFrame): DataFrame to search in.
    - target_value (str): Target value to search for.


    Returns:
    - list of int: List of row numbers where the target value is located.
    """
    rows = []
    for index, value in enumerate(
        df.iloc[:, 1], start=1
    ):  # Using positional indexing for the second column
        if value == target_value:
            rows.append(index)
    return rows


def find_occurrences_of_number(df, number):
    """
    Finds occurrences of a specified number [Months] in the second column of the DataFrame.

    Args:
    - df (DataFrame): DataFrame to search in.
    - number (int): Number to search for.

    Returns:
    - list of int: List of row numbers where the specified number is found in the second column.
    """
    rows = []
    for index, value in enumerate(
        df.iloc[:, 2], start=1
    ):  # Using positional indexing for the second column
        if pd.isna(value):
            continue  # Skip NaN values
        try:
            if int(value) == number:
                rows.append(index)
        except ValueError:
            continue  # Skip non-numeric values
    return rows


def find_start_stop_indices(rows):
    """
    Finds the lowest and highest row numbers from the given list of row numbers. This allows you to know where the program codes begin and end

    Args:
    - rows (list of int): List of row numbers.

    Returns:
    - int: Lowest row number.
    - int: Highest row number.
    """
    if not rows:
        return None, None
    return min(rows), max(rows)


def check_occurrences_and_create_fields(number_occurrences, target_indices, df):
    """
    Check occurrences [Months] and create fields based on their values.

    Args:
    - number_occurrences (dict): Dictionary containing row numbers where each number is found.
    - target_indices (dict): Dictionary containing start and stop indices for each target program.
    - df (DataFrame): DataFrame containing the data.

    Returns:
    - dict: Dictionary containing created fields.
    """
    created_fields = {}
    for number, occurrences in number_occurrences.items():
        for row_number in occurrences:
            for program, indices in target_indices.items():
                if indices["start"] is not None and indices["stop"] is not None:
                    if indices["start"] <= row_number <= indices["stop"]:
                        col_value = df.iloc[
                            row_number - 1, 4
                        ]  # Assuming you want to check the column +1 over
                        Month_value = df.iloc[
                            row_number - 1, 2
                        ]  # Assuming you want to check the column +1 over
                        APA_value = df.iloc[row_number - 1, 35]
                        field_name = f"{program}_Month_{Month_value}_{col_value}: "
                        created_fields[field_name] = APA_value
    return created_fields


def write_to_excel(file_path, sheet_name, cell, value):
    """
    Writes a value to a specific cell in an Excel file.

    Args:
    - file_path (str): Path to the Excel file.
    - sheet_name (str): Name of the sheet to write to.
    - cell (str): Cell reference (e.g., 'E10').
    - value: Value to write to the cell.
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    sheet[cell] = value
    wb.save(file_path)


def write_program_values_to_excel(created_fields, excel_output_path, excel_sheet_name):
    """
    Write program values to the Excel spreadsheet.

    Args:
    - created_fields (dict): Dictionary containing program and month combinations as keys and their values.
    - excel_output_path (str): Path to the Excel output file.
    - excel_sheet_name (str): Name of the Excel sheet.

    Returns:
    - None
    """
    # Define the starting row and column for writing data
    start_row = 9  # Starting from row 9
    start_col = 5  # Starting from column E

    print("Writing data to Excel...")

    start_time = time.time()  # Record the start time

    # Initialize tqdm progress bar
    progress_bar = tqdm(total=len(created_fields), desc="Progress")

    # Loop through each program and its value
    for idx, (program, value) in enumerate(created_fields.items()):
        # Calculate the cell coordinates based on the index
        row = start_row + idx
        col = start_col + (
            idx % 4
        )  # Move one column over every 4 programs (assuming 4 months per program)

        # Simulate processing time (remove this line if not needed)
        time.sleep(1)  # Wait for 1 second (adjust as needed)

        # Write the value to the Excel spreadsheet
        write_to_excel(
            excel_output_path,
            excel_sheet_name,
            f"{openpyxl.utils.get_column_letter(col)}{row}",
            value,
        )

        # Update progress bar
        progress_bar.update(1)
        progress_bar.set_postfix(
            {"Program": program}
        )  # Update progress bar description with the current program

    progress_bar.close()

    end_time = time.time()  # Record the end time
    elapsed_time = end_time - start_time  # Calculate the elapsed time

    print(f"Data written successfully to Excel in {elapsed_time:.2f} seconds.")


def batch_load_values(created_fields, excel_output_path, excel_sheet_name):
   
    # Define the list of cell references and values to be written
    cell_value_list = [
        # Program C Placements
        ("E58", created_fields.get("Prog_C_Month_1_TK-3: ", 0)),
        ("E59", created_fields.get("Prog_C_Month_1_4-6: ", 0)),
        ("E60", created_fields.get("Prog_C_Month_1_7-8: ", 0)),
        ("E61", created_fields.get("Prog_C_Month_1_9-12: ", 0)),
        ("E57", created_fields.get("Prog_C_TK_Month_1_TK-3: ", 0)),
        ("F58", created_fields.get("Prog_C_Month_2_TK-3: ", 0)),
        ("F59", created_fields.get("Prog_C_Month_2_4-6: ", 0)),
        ("F60", created_fields.get("Prog_C_Month_2_7-8: ", 0)),
        ("F61", created_fields.get("Prog_C_Month_2_9-12: ", 0)),
        ("F57", created_fields.get("Prog_C_TK_Month_2_TK-3: ", 0)),
        ("G58", created_fields.get("Prog_C_Month_3_TK-3: ", 0)),
        ("G59", created_fields.get("Prog_C_Month_3_4-6: ", 0)),
        ("G60", created_fields.get("Prog_C_Month_3_7-8: ", 0)),
        ("G61", created_fields.get("Prog_C_Month_3_9-12: ", 0)),
        ("G57", created_fields.get("Prog_C_TK_Month_3_TK-3: ", 0)),
        ("H58", created_fields.get("Prog_C_Month_4_TK-3: ", 0)),
        ("H59", created_fields.get("Prog_C_Month_4_4-6: ", 0)),
        ("H60", created_fields.get("Prog_C_Month_4_7-8: ", 0)),
        ("H61", created_fields.get("Prog_C_Month_4_9-12: ", 0)),
        ("H57", created_fields.get("Prog_C_TK_Month_4_TK-3: ", 0)),
        ("I58", created_fields.get("Prog_C_Month_5_TK-3: ", 0)),
        ("I59", created_fields.get("Prog_C_Month_5_4-6: ", 0)),
        ("I60", created_fields.get("Prog_C_Month_5_7-8: ", 0)),
        ("I61", created_fields.get("Prog_C_Month_5_9-12: ", 0)),
        ("I57", created_fields.get("Prog_C_TK_Month_5_TK-3: ", 0)),
        ("J58", created_fields.get("Prog_C_Month_6_TK-3: ", 0)),
        ("J59", created_fields.get("Prog_C_Month_6_4-6: ", 0)),
        ("J60", created_fields.get("Prog_C_Month_6_7-8: ", 0)),
        ("J61", created_fields.get("Prog_C_Month_6_9-12: ", 0)),
        ("J57", created_fields.get("Prog_C_TK_Month_6_TK-3: ", 0)),
        ("K58", created_fields.get("Prog_C_Month_7_TK-3: ", 0)),
        ("K59", created_fields.get("Prog_C_Month_7_4-6: ", 0)),
        ("K60", created_fields.get("Prog_C_Month_7_7-8: ", 0)),
        ("K61", created_fields.get("Prog_C_Month_7_9-12: ", 0)),
        ("K57", created_fields.get("Prog_C_TK_Month_7_TK-3: ", 0)),
        ("L58", created_fields.get("Prog_C_Month_8_TK-3: ", 0)),
        ("L59", created_fields.get("Prog_C_Month_8_4-6: ", 0)),
        ("L60", created_fields.get("Prog_C_Month_8_7-8: ", 0)),
        ("L61", created_fields.get("Prog_C_Month_8_9-12: ", 0)),
        ("L57", created_fields.get("Prog_C_TK_Month_8_TK-3: ", 0)),
        ("M58", created_fields.get("Prog_C_Month_9_TK-3: ", 0)),
        ("M59", created_fields.get("Prog_C_Month_9_4-6: ", 0)),
        ("M60", created_fields.get("Prog_C_Month_9_7-8: ", 0)),
        ("M61", created_fields.get("Prog_C_Month_9_9-12: ", 0)),
        ("M57", created_fields.get("Prog_C_TK_Month_9_TK-3: ", 0)),
        ("N58", created_fields.get("Prog_C_Month_10_TK-3: ", 0)),
        ("N59", created_fields.get("Prog_C_Month_10_4-6: ", 0)),
        ("N60", created_fields.get("Prog_C_Month_10_7-8: ", 0)),
        ("N61", created_fields.get("Prog_C_Month_10_9-12: ", 0)),
        ("N57", created_fields.get("Prog_C_TK_Month_10_TK-3: ", 0)),
        ("O58", created_fields.get("Prog_C_Month_11_TK-3: ", 0)),
        ("O59", created_fields.get("Prog_C_Month_11_4-6: ", 0)),
        ("O60", created_fields.get("Prog_C_Month_11_7-8: ", 0)),
        ("O61", created_fields.get("Prog_C_Month_11_9-12: ", 0)),
        ("O57", created_fields.get("Prog_C_TK_Month_11_TK-3: ", 0)),
        ("P58", created_fields.get("Prog_C_Month_12_TK-3: ", 0)),
        ("P59", created_fields.get("Prog_C_Month_12_4-6: ", 0)),
        ("P60", created_fields.get("Prog_C_Month_12_7-8: ", 0)),
        ("P61", created_fields.get("Prog_C_Month_12_9-12: ", 0)),
        ("P57", created_fields.get("Prog_C_TK_Month_12_TK-3: ", 0)),
        # Program N Placements
        ("E83", created_fields.get("Prog_N_Month_1_TK-3: ", 0)),
        ("E84", created_fields.get("Prog_N_Month_1_4-6: ", 0)),
        ("E85", created_fields.get("Prog_N_Month_1_7-8: ", 0)),
        ("E86", created_fields.get("Prog_N_Month_1_9-12: ", 0)),
        ("E82", created_fields.get("Prog_N_TK_Month_1_TK-3: ", 0)),
        ("F83", created_fields.get("Prog_N_Month_2_TK-3: ", 0)),
        ("F84", created_fields.get("Prog_N_Month_2_4-6: ", 0)),
        ("F85", created_fields.get("Prog_N_Month_2_7-8: ", 0)),
        ("F86", created_fields.get("Prog_N_Month_2_9-12: ", 0)),
        ("F82", created_fields.get("Prog_N_TK_Month_2_TK-3: ", 0)),
        ("G83", created_fields.get("Prog_N_Month_3_TK-3: ", 0)),
        ("G84", created_fields.get("Prog_N_Month_3_4-6: ", 0)),
        ("G85", created_fields.get("Prog_N_Month_3_7-8: ", 0)),
        ("G86", created_fields.get("Prog_N_Month_3_9-12: ", 0)),
        ("G82", created_fields.get("Prog_N_TK_Month_3_TK-3: ", 0)),
        ("H83", created_fields.get("Prog_N_Month_4_TK-3: ", 0)),
        ("H84", created_fields.get("Prog_N_Month_4_4-6: ", 0)),
        ("H85", created_fields.get("Prog_N_Month_4_7-8: ", 0)),
        ("H86", created_fields.get("Prog_N_Month_4_9-12: ", 0)),
        ("H82", created_fields.get("Prog_N_TK_Month_4_TK-3: ", 0)),
        ("I83", created_fields.get("Prog_N_Month_5_TK-3: ", 0)),
        ("I84", created_fields.get("Prog_N_Month_5_4-6: ", 0)),
        ("I85", created_fields.get("Prog_N_Month_5_7-8: ", 0)),
        ("I86", created_fields.get("Prog_N_Month_5_9-12: ", 0)),
        ("I82", created_fields.get("Prog_N_TK_Month_5_TK-3: ", 0)),
        ("J83", created_fields.get("Prog_N_Month_6_TK-3: ", 0)),
        ("J84", created_fields.get("Prog_N_Month_6_4-6: ", 0)),
        ("J85", created_fields.get("Prog_N_Month_6_7-8: ", 0)),
        ("J86", created_fields.get("Prog_N_Month_6_9-12: ", 0)),
        ("J82", created_fields.get("Prog_N_TK_Month_6_TK-3: ", 0)),
        ("K83", created_fields.get("Prog_N_Month_7_TK-3: ", 0)),
        ("K84", created_fields.get("Prog_N_Month_7_4-6: ", 0)),
        ("K85", created_fields.get("Prog_N_Month_7_7-8: ", 0)),
        ("K86", created_fields.get("Prog_N_Month_7_9-12: ", 0)),
        ("K82", created_fields.get("Prog_N_TK_Month_7_TK-3: ", 0)),
        ("L83", created_fields.get("Prog_N_Month_8_TK-3: ", 0)),
        ("L84", created_fields.get("Prog_N_Month_8_4-6: ", 0)),
        ("L85", created_fields.get("Prog_N_Month_8_7-8: ", 0)),
        ("L86", created_fields.get("Prog_N_Month_8_9-12: ", 0)),
        ("L82", created_fields.get("Prog_N_TK_Month_8_TK-3: ", 0)),
        ("M83", created_fields.get("Prog_N_Month_9_TK-3: ", 0)),
        ("M84", created_fields.get("Prog_N_Month_9_4-6: ", 0)),
        ("M85", created_fields.get("Prog_N_Month_9_7-8: ", 0)),
        ("M86", created_fields.get("Prog_N_Month_9_9-12: ", 0)),
        ("M82", created_fields.get("Prog_N_TK_Month_9_TK-3: ", 0)),
        ("N83", created_fields.get("Prog_N_Month_10_TK-3: ", 0)),
        ("N84", created_fields.get("Prog_N_Month_10_4-6: ", 0)),
        ("N85", created_fields.get("Prog_N_Month_10_7-8: ", 0)),
        ("N86", created_fields.get("Prog_N_Month_10_9-12: ", 0)),
        ("N82", created_fields.get("Prog_N_TK_Month_10_TK-3: ", 0)),
        ("O83", created_fields.get("Prog_N_Month_11_TK-3: ", 0)),
        ("O84", created_fields.get("Prog_N_Month_11_4-6: ", 0)),
        ("O85", created_fields.get("Prog_N_Month_11_7-8: ", 0)),
        ("O86", created_fields.get("Prog_N_Month_11_9-12: ", 0)),
        ("O82", created_fields.get("Prog_N_TK_Month_11_TK-3: ", 0)),
        ("P83", created_fields.get("Prog_N_Month_12_TK-3: ", 0)),
        ("P84", created_fields.get("Prog_N_Month_12_4-6: ", 0)),
        ("P85", created_fields.get("Prog_N_Month_12_7-8: ", 0)),
        ("P86", created_fields.get("Prog_N_Month_12_9-12: ", 0)),
        ("P82", created_fields.get("Prog_N_TK_Month_12_TK-3: ", 0)),
        # Program J Placements
        ("E65", created_fields.get("Prog_J_Month_1_TK-3: ", 0)),
        ("E66", created_fields.get("Prog_J_Month_1_4-6: ", 0)),
        ("E67", created_fields.get("Prog_J_Month_1_7-8: ", 0)),
        ("E68", created_fields.get("Prog_J_Month_1_9-12: ", 0)),
        ("E64", created_fields.get("Prog_J_TK_Month_1_TK-3: ", 0)),
        ("F65", created_fields.get("Prog_J_Month_2_TK-3: ", 0)),
        ("F66", created_fields.get("Prog_J_Month_2_4-6: ", 0)),
        ("F67", created_fields.get("Prog_J_Month_2_7-8: ", 0)),
        ("F68", created_fields.get("Prog_J_Month_2_9-12: ", 0)),
        ("F64", created_fields.get("Prog_J_TK_Month_2_TK-3: ", 0)),
        ("G65", created_fields.get("Prog_J_Month_3_TK-3: ", 0)),
        ("G66", created_fields.get("Prog_J_Month_3_4-6: ", 0)),
        ("G67", created_fields.get("Prog_J_Month_3_7-8: ", 0)),
        ("G68", created_fields.get("Prog_J_Month_3_9-12: ", 0)),
        ("G64", created_fields.get("Prog_J_TK_Month_3_TK-3: ", 0)),
        ("H65", created_fields.get("Prog_J_Month_4_TK-3: ", 0)),
        ("H66", created_fields.get("Prog_J_Month_4_4-6: ", 0)),
        ("H67", created_fields.get("Prog_J_Month_4_7-8: ", 0)),
        ("H68", created_fields.get("Prog_J_Month_4_9-12: ", 0)),
        ("H64", created_fields.get("Prog_J_TK_Month_4_TK-3: ", 0)),
        ("I65", created_fields.get("Prog_J_Month_5_TK-3: ", 0)),
        ("I66", created_fields.get("Prog_J_Month_5_4-6: ", 0)),
        ("I67", created_fields.get("Prog_J_Month_5_7-8: ", 0)),
        ("I68", created_fields.get("Prog_J_Month_5_9-12: ", 0)),
        ("I64", created_fields.get("Prog_J_TK_Month_5_TK-3: ", 0)),
        ("J65", created_fields.get("Prog_J_Month_6_TK-3: ", 0)),
        ("J66", created_fields.get("Prog_J_Month_6_4-6: ", 0)),
        ("J67", created_fields.get("Prog_J_Month_6_7-8: ", 0)),
        ("J68", created_fields.get("Prog_J_Month_6_9-12: ", 0)),
        ("J64", created_fields.get("Prog_J_TK_Month_6_TK-3: ", 0)),
        ("K65", created_fields.get("Prog_J_Month_7_TK-3: ", 0)),
        ("K66", created_fields.get("Prog_J_Month_7_4-6: ", 0)),
        ("K67", created_fields.get("Prog_J_Month_7_7-8: ", 0)),
        ("K68", created_fields.get("Prog_J_Month_7_9-12: ", 0)),
        ("K64", created_fields.get("Prog_J_TK_Month_7_TK-3: ", 0)),
        ("L65", created_fields.get("Prog_J_Month_8_TK-3: ", 0)),
        ("L66", created_fields.get("Prog_J_Month_8_4-6: ", 0)),
        ("L67", created_fields.get("Prog_J_Month_8_7-8: ", 0)),
        ("L68", created_fields.get("Prog_J_Month_8_9-12: ", 0)),
        ("L64", created_fields.get("Prog_J_TK_Month_8_TK-3: ", 0)),
        ("M65", created_fields.get("Prog_J_Month_9_TK-3: ", 0)),
        ("M66", created_fields.get("Prog_J_Month_9_4-6: ", 0)),
        ("M67", created_fields.get("Prog_J_Month_9_7-8: ", 0)),
        ("M68", created_fields.get("Prog_J_Month_9_9-12: ", 0)),
        ("M64", created_fields.get("Prog_J_TK_Month_9_TK-3: ", 0)),
        ("N65", created_fields.get("Prog_J_Month_10_TK-3: ", 0)),
        ("N66", created_fields.get("Prog_J_Month_10_4-6: ", 0)),
        ("N67", created_fields.get("Prog_J_Month_10_7-8: ", 0)),
        ("N68", created_fields.get("Prog_J_Month_10_9-12: ", 0)),
        ("N64", created_fields.get("Prog_J_TK_Month_10_TK-3: ", 0)),
        ("O65", created_fields.get("Prog_J_Month_11_TK-3: ", 0)),
        ("O66", created_fields.get("Prog_J_Month_11_4-6: ", 0)),
        ("O67", created_fields.get("Prog_J_Month_11_7-8: ", 0)),
        ("O68", created_fields.get("Prog_J_Month_11_9-12: ", 0)),
        ("O64", created_fields.get("Prog_J_TK_Month_11_TK-3: ", 0)),
        ("P65", created_fields.get("Prog_J_Month_12_TK-3: ", 0)),
        ("P66", created_fields.get("Prog_J_Month_12_4-6: ", 0)),
        ("P67", created_fields.get("Prog_J_Month_12_7-8: ", 0)),
        ("P68", created_fields.get("Prog_J_Month_12_9-12: ", 0)),
        ("P64", created_fields.get("Prog_J_TK_Month_12_TK-3: ", 0)),
        # Program k Placements
        ("E83", created_fields.get("Prog_K_Month_1_TK-3: ", 0)),
        ("E84", created_fields.get("Prog_K_Month_1_4-6: ", 0)),
        ("E85", created_fields.get("Prog_K_Month_1_7-8: ", 0)),
        ("E86", created_fields.get("Prog_K_Month_1_9-12: ", 0)),
        ("E82", created_fields.get("Prog_K_TK_Month_1_TK-3: ", 0)),
        ("F83", created_fields.get("Prog_K_Month_2_TK-3: ", 0)),
        ("F84", created_fields.get("Prog_K_Month_2_4-6: ", 0)),
        ("F85", created_fields.get("Prog_K_Month_2_7-8: ", 0)),
        ("F86", created_fields.get("Prog_K_Month_2_9-12: ", 0)),
        ("F82", created_fields.get("Prog_K_TK_Month_2_TK-3: ", 0)),
        ("G83", created_fields.get("Prog_K_Month_3_TK-3: ", 0)),
        ("G84", created_fields.get("Prog_K_Month_3_4-6: ", 0)),
        ("G85", created_fields.get("Prog_K_Month_3_7-8: ", 0)),
        ("G86", created_fields.get("Prog_K_Month_3_9-12: ", 0)),
        ("G82", created_fields.get("Prog_K_TK_Month_3_TK-3: ", 0)),
        ("H83", created_fields.get("Prog_K_Month_4_TK-3: ", 0)),
        ("H84", created_fields.get("Prog_K_Month_4_4-6: ", 0)),
        ("H85", created_fields.get("Prog_K_Month_4_7-8: ", 0)),
        ("H86", created_fields.get("Prog_K_Month_4_9-12: ", 0)),
        ("H82", created_fields.get("Prog_K_TK_Month_4_TK-3: ", 0)),
        ("I83", created_fields.get("Prog_K_Month_5_TK-3: ", 0)),
        ("I84", created_fields.get("Prog_K_Month_5_4-6: ", 0)),
        ("I85", created_fields.get("Prog_K_Month_5_7-8: ", 0)),
        ("I86", created_fields.get("Prog_K_Month_5_9-12: ", 0)),
        ("I82", created_fields.get("Prog_K_TK_Month_5_TK-3: ", 0)),
        ("J83", created_fields.get("Prog_K_Month_6_TK-3: ", 0)),
        ("J84", created_fields.get("Prog_K_Month_6_4-6: ", 0)),
        ("J85", created_fields.get("Prog_K_Month_6_7-8: ", 0)),
        ("J86", created_fields.get("Prog_K_Month_6_9-12: ", 0)),
        ("J82", created_fields.get("Prog_K_TK_Month_6_TK-3: ", 0)),
        ("K83", created_fields.get("Prog_K_Month_7_TK-3: ", 0)),
        ("K84", created_fields.get("Prog_K_Month_7_4-6: ", 0)),
        ("K85", created_fields.get("Prog_K_Month_7_7-8: ", 0)),
        ("K86", created_fields.get("Prog_K_Month_7_9-12: ", 0)),
        ("K82", created_fields.get("Prog_K_TK_Month_7_TK-3: ", 0)),
        ("L83", created_fields.get("Prog_K_Month_8_TK-3: ", 0)),
        ("L84", created_fields.get("Prog_K_Month_8_4-6: ", 0)),
        ("L85", created_fields.get("Prog_K_Month_8_7-8: ", 0)),
        ("L86", created_fields.get("Prog_K_Month_8_9-12: ", 0)),
        ("L82", created_fields.get("Prog_K_TK_Month_8_TK-3: ", 0)),
        ("M83", created_fields.get("Prog_K_Month_9_TK-3: ", 0)),
        ("M84", created_fields.get("Prog_K_Month_9_4-6: ", 0)),
        ("M85", created_fields.get("Prog_K_Month_9_7-8: ", 0)),
        ("M86", created_fields.get("Prog_K_Month_9_9-12: ", 0)),
        ("M82", created_fields.get("Prog_K_TK_Month_9_TK-3: ", 0)),
        ("N83", created_fields.get("Prog_K_Month_10_TK-3: ", 0)),
        ("N84", created_fields.get("Prog_K_Month_10_4-6: ", 0)),
        ("N85", created_fields.get("Prog_K_Month_10_7-8: ", 0)),
        ("N86", created_fields.get("Prog_K_Month_10_9-12: ", 0)),
        ("N82", created_fields.get("Prog_K_TK_Month_10_TK-3: ", 0)),
        ("O83", created_fields.get("Prog_K_Month_11_TK-3: ", 0)),
        ("O84", created_fields.get("Prog_K_Month_11_4-6: ", 0)),
        ("O85", created_fields.get("Prog_K_Month_11_7-8: ", 0)),
        ("O86", created_fields.get("Prog_K_Month_11_9-12: ", 0)),
        ("O82", created_fields.get("Prog_K_TK_Month_11_TK-3: ", 0)),
        ("P83", created_fields.get("Prog_K_Month_12_TK-3: ", 0)),
        ("P84", created_fields.get("Prog_K_Month_12_4-6: ", 0)),
        ("P85", created_fields.get("Prog_K_Month_12_7-8: ", 0)),
        ("P86", created_fields.get("Prog_K_Month_12_9-12: ", 0)),
        ("P82", created_fields.get("Prog_K_TK_Month_12_TK-3: ", 0)),
    ]
    # Open the Excel file
    wb = openpyxl.load_workbook(excel_output_path)
    sheet = wb[excel_sheet_name]

    # Batch load values into cells
    for cell, value in cell_value_list:
        sheet[cell] = value

    # Save the changes
    wb.save(excel_output_path)


def main():
    """
    Main function to execute the program.
    """
    file_path = (
       "C:\\Users\\Shawn\\Downloads\\PrintMonthlyAttendanceSummaryTotals_20251009_134126_294513e.xlsx"
    )
    excel_output_path = "C:\\Users\\Shawn\\Downloads\\2025-2026_I4C_ADA_Reconciliation.xlsx"
    excel_sheet_name = "Template- Apportionment Summary"

    target_values = {
        "Program C Charter Resident": "Prog_C",
        "Program C Charter Resident -  Transitional Kindergarten(TK)": "Prog_C_TK",
        "Program N Non-Resident Charter": "Prog_N",
        "Program N Non-Resident Charter -  Transitional Kindergarten(TK)": "Prog_N_TK",
        "Program J Indep Study Charter Resident": "Prog_J",
        "Program J Indep Study Charter Non-Resident -  Transitional Kindergarten(TK)": "Prog_J_TK",
        "Program K Indep Study Charter Non-Resident": "Prog_K",
        "Program K Indep Study Charter Non-Resident -  Transitional Kindergarten(TK)": "Prog_K_TK",
    }

    # Read the Excel file into a DataFrame
    df = pd.read_excel(
        file_path, header=None
    )  # Specify header=None to indicate no header

    # Initialize dictionaries to store start and stop indices for each target value
    target_indices = {
        value: {"start": None, "stop": None} for value in target_values.values()
    }

    # Find row numbers for each target value and update dictionaries with start and stop indices
    for target_value, dict_name in target_values.items():
        rows = find_row_with_value(df, target_value)
        start, stop = find_start_stop_indices(rows)
        target_indices[dict_name]["start"] = start
        target_indices[dict_name]["stop"] = stop

    prog_c_start = target_indices["Prog_C_TK"]["start"]
    prog_c_tk_stop = target_indices["Prog_N"]["start"]

    # Update the stop index of Prog_C to be one less than the start index of Prog_C_TK if start is not none
    if prog_c_start is not None and prog_c_tk_stop is not None:
        target_indices["Prog_C"]["stop"] = prog_c_start - 1

    # Update the stop index of Prog_C_TK to be one less than the stop index of Prog_C
    if prog_c_tk_stop is not None:
        target_indices["Prog_C_TK"]["stop"] = prog_c_tk_stop - 1

    # Update the stop index of Prog_N to be one less than the start index of Prog_N_TK if Prog_N_TK is not None
    prog_n_tk_start = target_indices["Prog_N_TK"]["start"]
    if prog_n_tk_start is not None:
        target_indices["Prog_N"]["stop"] = prog_n_tk_start - 1

    # Update the stop index of Prog_N_TK, Prog_J, and Prog_K to be one less than the start index of the next program if start is not none
    programs = ["Prog_N_TK", "Prog_J", "Prog_K"]
    for i in range(len(programs) - 1):
        start = target_indices[programs[i]]["start"]
        next_start = target_indices[programs[i + 1]]["start"]
        if start is not None and next_start is not None:
            target_indices[programs[i]]["stop"] = next_start - 1

    # Display start and stop indices of each program
    print("Current start and stop indices of programs:")
    for program, indices in target_indices.items():
        start = indices.get("start", "Not defined")
        stop = indices.get("stop", "Not defined")
        print(f"Program: {program}, Start: {start}, Stop: {stop}")

    # Allow the user to verify and update start and stop indices if necessary
    for program in target_indices.keys():
        confirm = input(
            f"Are the saved start and stop indices for {program} correct? (yes/no): "
        ).lower()
        if confirm == "no":
            while True:
                user_input = input(
                    f"Enter new start and stop indices for {program} separated by a comma (e.g., 'start, stop'): "
                )
                new_indices = user_input.split(",")

                # Validate user input and update target_indices
                if len(new_indices) == 2:
                    start = (
                        int(new_indices[0].strip())
                        if new_indices[0].strip().lower() != "none"
                        else None
                    )
                    stop = (
                        int(new_indices[1].strip())
                        if new_indices[1].strip().lower() != "none"
                        else None
                    )
                    target_indices[program]["start"] = start
                    target_indices[program]["stop"] = stop
                    break  # Exit the loop if input is valid
                else:
                    print(
                        "Invalid input. Please enter valid integer start and stop indices separated by a comma."
                    )

    # Print out the dictionaries containing start and stop indices for each target value
    for target_value, indices in target_indices.items():
        print(f"{target_value}: {indices}")

    # Find occurrences of numbers 1 through 12 in the second column of the DataFrame
    number_occurrences = {}
    for i in range(1, 13):
        occurrences = find_occurrences_of_number(df, i)
        number_occurrences[i] = occurrences
        print(f"Occurrences of Month '{i}' in rows:")
        print(occurrences)

    created_fields = check_occurrences_and_create_fields(
        number_occurrences, target_indices, df
    )
    for key, value in created_fields.items():
        print(f"{key}: {value}")

    # Call batch_load_values to batch load the values into the Excel sheet
    batch_load_values(created_fields, excel_output_path, excel_sheet_name)

    """
    # Write the value of 'Prog_C_Month_1_TK' to cell E10 in the spreadsheet
    prog_c_month_1_tk_value = created_fields.get("Prog_C_Month_1_TK-3: ", 0)  # Assuming default value is 0 if key is not found
    write_to_excel(excel_output_path, excel_sheet_name, 'E10', prog_c_month_1_tk_value)

    # Write the value of 'Prog_C_Month_1_4-6' to cell E10 in the spreadsheet
    prog_C_Month_1_4_6 = created_fields.get("Prog_C_Month_1_4-6: ", 0)  # Assuming default value is 0 if key is not found
    write_to_excel(excel_output_path, excel_sheet_name, 'E11', prog_C_Month_1_4_6)

    # Write the value of 'Prog_C_Month_1_7-8' to cell E10 in the spreadsheet
    prog_C_Month_1_7_8 = created_fields.get("Prog_C_Month_1_7-8: ", 0)  # Assuming default value is 0 if key is not found
    write_to_excel(excel_output_path, excel_sheet_name, 'E12', prog_C_Month_1_7_8)

    # Write the value of 'Prog_C_Month_1_4-6' to cell E10 in the spreadsheet
    prog_C_Month_1_9_12 = created_fields.get("Prog_C_Month_1_9-12: ", 0)  # Assuming default value is 0 if key is not found
    write_to_excel(excel_output_path, excel_sheet_name, 'E13', prog_C_Month_1_9_12)

    # Write the value of 'Prog_C_Month_1_4-6' to cell E10 in the spreadsheet
    prog_C_Month_1_TK = created_fields.get("Prog_C_TK_Month_1_TK-3: ", 0)  # Assuming default value is 0 if key is not found
    write_to_excel(excel_output_path, excel_sheet_name, 'E9', prog_C_Month_1_TK)

    # Write the value of 'Prog_C_Month_1_TK' to cell E10 in the spreadsheet
    prog_c_month_2_tk_value = created_fields.get("Prog_C_Month_2_TK-3: ", 0)  # Assuming default value is 0 if key is not found
    write_to_excel(excel_output_path, excel_sheet_name, 'F10', prog_c_month_2_tk_value)

    # Write the value of 'Prog_C_Month_1_4-6' to cell E10 in the spreadsheet
    prog_C_Month_2_4_6 = created_fields.get("Prog_C_Month_2_4-6: ", 0)  # Assuming default value is 0 if key is not found
    write_to_excel(excel_output_path, excel_sheet_name, 'F11', prog_C_Month_2_4_6)

    # Write the value of 'Prog_C_Month_1_7-8' to cell E10 in the spreadsheet
    prog_C_Month_2_7_8 = created_fields.get("Prog_C_Month_2_7-8: ", 0)  # Assuming default value is 0 if key is not found
    write_to_excel(excel_output_path, excel_sheet_name, 'F12', prog_C_Month_2_7_8)

    # Write the value of 'Prog_C_Month_1_4-6' to cell E10 in the spreadsheet
    prog_C_Month_2_9_12 = created_fields.get("Prog_C_Month_2_9-12: ", 0)  # Assuming default value is 0 if key is not found
    write_to_excel(excel_output_path, excel_sheet_name, 'F13', prog_C_Month_2_9_12)

    # Write the value of 'Prog_C_Month_1_4-6' to cell E10 in the spreadsheet
    prog_C_Month_2_TK = created_fields.get("Prog_C_TK_Month_2_TK-3: ", 0)  # Assuming default value is 0 if key is not found
    write_to_excel(excel_output_path, excel_sheet_name, 'F9', prog_C_Month_2_TK)

if __name__ == "__main__":

    main()
    """


if __name__ == "__main__":
    main()
