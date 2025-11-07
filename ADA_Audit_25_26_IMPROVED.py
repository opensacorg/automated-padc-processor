import pandas as pd
import openpyxl
from tqdm import tqdm
import time


# =============================================================================
# UTILITY FUNCTIONS FOR SEARCHING AND PROCESSING DATA
# =============================================================================

def find_rows_containing_program_name(student_data, program_name_to_find):
    """
    Searches through student data to find all rows that contain a specific program name.
    
    What this does:
    - Looks through the second column of data (column B in Excel terms)
    - Finds every row where the program name appears
    - Returns a list of row numbers where it was found
    
    Parameters:
    - student_data: The Excel data loaded as a DataFrame
    - program_name_to_find: The exact program name we're looking for (like "Program C Charter Resident")
    
    Returns:
    - A list of row numbers (starting from 1, like Excel row numbers)
    """
    matching_row_numbers = []
    
    # Look through each cell in the second column (index 1 = column B)
    for row_index, cell_value in enumerate(student_data.iloc[:, 1], start=1):
        if cell_value == program_name_to_find:
            matching_row_numbers.append(row_index)
    
    return matching_row_numbers


def find_rows_containing_month_number(student_data, month_number_to_find):
    """
    Searches for rows that contain a specific month number.
    
    What this does:
    - Looks through the third column of data (column C in Excel terms)
    - Finds every row where the month number appears (1=January, 2=February, etc.)
    - Returns a list of row numbers where it was found
    
    Parameters:
    - student_data: The Excel data loaded as a DataFrame
    - month_number_to_find: The month number to search for (1-12)
    
    Returns:
    - A list of row numbers where this month number was found
    """
    matching_row_numbers = []
    
    # Look through each cell in the third column (index 2 = column C)
    for row_index, cell_value in enumerate(student_data.iloc[:, 2], start=1):
        # Skip empty cells
        if pd.isna(cell_value):
            continue
        
        # Try to convert to number and compare
        try:
            if int(cell_value) == month_number_to_find:
                matching_row_numbers.append(row_index)
        except ValueError:
            # Skip cells that aren't numbers
            continue
    
    return matching_row_numbers


def find_program_boundary_rows(list_of_row_numbers):
    """
    Finds where a program's data starts and ends in the spreadsheet.
    
    What this does:
    - Takes a list of row numbers where a program appears
    - Finds the first row (lowest number) and last row (highest number)
    - This tells us the "boundaries" of where this program's data is located
    
    Parameters:
    - list_of_row_numbers: List of row numbers where a program was found
    
    Returns:
    - Two numbers: (first_row, last_row) or (None, None) if no rows found
    """
    if not list_of_row_numbers:
        return None, None
    
    first_row = min(list_of_row_numbers)
    last_row = max(list_of_row_numbers)
    
    return first_row, last_row


def extract_student_attendance_data(monthly_attendance_by_program, program_boundary_info, student_data):
    """
    Extracts attendance data for each program and month combination.
    
    What this does:
    - For each month and program combination, it creates a descriptive name
    - It extracts the attendance value from the appropriate column
    - Creates a dictionary with all the attendance data organized by program and month
    
    Parameters:
    - monthly_attendance_by_program: Dict showing which months appear in which rows
    - program_boundary_info: Dict showing where each program's data starts and ends
    - student_data: The Excel data loaded as a DataFrame
    
    Returns:
    - A dictionary with descriptive names as keys and attendance values as values
    """
    attendance_data_dictionary = {}
    
    # For each month (1-12)
    for month_number, rows_with_this_month in monthly_attendance_by_program.items():
        # For each row where this month appears
        for current_row_number in rows_with_this_month:
            # For each program we're tracking
            for program_code, boundary_info in program_boundary_info.items():
                # Check if we have valid boundaries for this program
                program_start_row = boundary_info["start"]
                program_end_row = boundary_info["stop"]
                
                if program_start_row is not None and program_end_row is not None:
                    # Check if this row falls within this program's boundaries
                    if program_start_row <= current_row_number <= program_end_row:
                        # Extract the age group from column E (index 4)
                        age_group = student_data.iloc[current_row_number - 1, 4]
                        
                        # Extract the month number from column C (index 2)
                        month_value = student_data.iloc[current_row_number - 1, 2]
                        
                        # Extract the attendance value from column AJ (index 35)
                        attendance_value = student_data.iloc[current_row_number - 1, 35]
                        
                        # Create a descriptive name for this data point
                        descriptive_field_name = f"{program_code}_Month_{month_value}_{age_group}: "
                        
                        # Store the attendance value with this descriptive name
                        attendance_data_dictionary[descriptive_field_name] = attendance_value
    
    return attendance_data_dictionary


# =============================================================================
# EXCEL WRITING FUNCTIONS
# =============================================================================

def write_single_value_to_excel_cell(excel_file_path, worksheet_name, target_cell, value_to_write):
    """
    Writes a single value to a specific cell in an Excel file.
    
    What this does:
    - Opens an Excel file
    - Finds the specified worksheet
    - Writes a value to a specific cell (like "E10")
    - Saves the file
    
    Parameters:
    - excel_file_path: Full path to the Excel file
    - worksheet_name: Name of the worksheet tab
    - target_cell: Cell reference like "E10", "F15", etc.
    - value_to_write: The value to put in that cell
    """
    workbook = openpyxl.load_workbook(excel_file_path)
    worksheet = workbook[worksheet_name]
    worksheet[target_cell] = value_to_write
    workbook.save(excel_file_path)


def write_program_data_one_by_one(attendance_data, output_excel_path, output_worksheet_name):
    """
    Writes attendance data to Excel one cell at a time (slower method).
    
    This function is currently not used but kept for reference.
    It writes each value individually, which is slower than batch writing.
    
    Parameters:
    - attendance_data: Dictionary with attendance data
    - output_excel_path: Path to the output Excel file
    - output_worksheet_name: Name of the worksheet to write to
    """
    # Define where to start writing data in the Excel sheet
    starting_row = 9
    starting_column = 5  # Column E
    
    print("Writing data to Excel one cell at a time...")
    
    start_time = time.time()
    progress_bar = tqdm(total=len(attendance_data), desc="Writing Progress")
    
    # Write each piece of data one by one
    for index, (data_description, attendance_value) in enumerate(attendance_data.items()):
        # Calculate where to put this data
        current_row = starting_row + index
        current_column = starting_column + (index % 4)  # Move columns every 4 entries
        
        # Add a small delay (remove if not needed)
        time.sleep(1)
        
        # Convert column number to letter and write the value
        cell_reference = f"{openpyxl.utils.get_column_letter(current_column)}{current_row}"
        write_single_value_to_excel_cell(
            output_excel_path,
            output_worksheet_name,
            cell_reference,
            attendance_value
        )
        
        # Update the progress bar
        progress_bar.update(1)
        progress_bar.set_postfix({"Current Program": data_description})
    
    progress_bar.close()
    
    end_time = time.time()
    total_time = end_time - start_time
    print(f"Data written successfully in {total_time:.2f} seconds.")


def write_all_attendance_data_to_excel_efficiently(attendance_data, output_excel_path, output_worksheet_name):
    """
    Efficiently writes all attendance data to Excel in one batch operation.
    
    What this does:
    - Opens the Excel file once
    - Maps each piece of attendance data to its correct cell location
    - Writes all data at once (much faster than writing one cell at a time)
    - Saves the file once
    
    IMPORTANT: This function has cell mapping conflicts that need to be fixed!
    Programs N and K are trying to write to the same cells, causing overwrites.
    
    Parameters:
    - attendance_data: Dictionary containing all the attendance data
    - output_excel_path: Path to the Excel file where data should be written
    - output_worksheet_name: Name of the worksheet tab to write to
    """
    
    # Create a master list mapping each data field to its target cell
    # IMPORTANT: This matches EXACTLY the same cells as the original cell_value_list
    cell_mapping_list = [
        # =================================================================
        # PROGRAM C CHARTER RESIDENT PLACEMENTS (Rows 57-61)
        # =================================================================
        ("E58", attendance_data.get("Prog_C_Month_1_TK-3: ", 0)),
        ("E59", attendance_data.get("Prog_C_Month_1_4-6: ", 0)),
        ("E60", attendance_data.get("Prog_C_Month_1_7-8: ", 0)),
        ("E61", attendance_data.get("Prog_C_Month_1_9-12: ", 0)),
        ("E57", attendance_data.get("Prog_C_TK_Month_1_TK-3: ", 0)),
        ("F58", attendance_data.get("Prog_C_Month_2_TK-3: ", 0)),
        ("F59", attendance_data.get("Prog_C_Month_2_4-6: ", 0)),
        ("F60", attendance_data.get("Prog_C_Month_2_7-8: ", 0)),
        ("F61", attendance_data.get("Prog_C_Month_2_9-12: ", 0)),
        ("F57", attendance_data.get("Prog_C_TK_Month_2_TK-3: ", 0)),
        ("G58", attendance_data.get("Prog_C_Month_3_TK-3: ", 0)),
        ("G59", attendance_data.get("Prog_C_Month_3_4-6: ", 0)),
        ("G60", attendance_data.get("Prog_C_Month_3_7-8: ", 0)),
        ("G61", attendance_data.get("Prog_C_Month_3_9-12: ", 0)),
        ("G57", attendance_data.get("Prog_C_TK_Month_3_TK-3: ", 0)),
        ("H58", attendance_data.get("Prog_C_Month_4_TK-3: ", 0)),
        ("H59", attendance_data.get("Prog_C_Month_4_4-6: ", 0)),
        ("H60", attendance_data.get("Prog_C_Month_4_7-8: ", 0)),
        ("H61", attendance_data.get("Prog_C_Month_4_9-12: ", 0)),
        ("H57", attendance_data.get("Prog_C_TK_Month_4_TK-3: ", 0)),
        ("I58", attendance_data.get("Prog_C_Month_5_TK-3: ", 0)),
        ("I59", attendance_data.get("Prog_C_Month_5_4-6: ", 0)),
        ("I60", attendance_data.get("Prog_C_Month_5_7-8: ", 0)),
        ("I61", attendance_data.get("Prog_C_Month_5_9-12: ", 0)),
        ("I57", attendance_data.get("Prog_C_TK_Month_5_TK-3: ", 0)),
        ("J58", attendance_data.get("Prog_C_Month_6_TK-3: ", 0)),
        ("J59", attendance_data.get("Prog_C_Month_6_4-6: ", 0)),
        ("J60", attendance_data.get("Prog_C_Month_6_7-8: ", 0)),
        ("J61", attendance_data.get("Prog_C_Month_6_9-12: ", 0)),
        ("J57", attendance_data.get("Prog_C_TK_Month_6_TK-3: ", 0)),
        ("K58", attendance_data.get("Prog_C_Month_7_TK-3: ", 0)),
        ("K59", attendance_data.get("Prog_C_Month_7_4-6: ", 0)),
        ("K60", attendance_data.get("Prog_C_Month_7_7-8: ", 0)),
        ("K61", attendance_data.get("Prog_C_Month_7_9-12: ", 0)),
        ("K57", attendance_data.get("Prog_C_TK_Month_7_TK-3: ", 0)),
        ("L58", attendance_data.get("Prog_C_Month_8_TK-3: ", 0)),
        ("L59", attendance_data.get("Prog_C_Month_8_4-6: ", 0)),
        ("L60", attendance_data.get("Prog_C_Month_8_7-8: ", 0)),
        ("L61", attendance_data.get("Prog_C_Month_8_9-12: ", 0)),
        ("L57", attendance_data.get("Prog_C_TK_Month_8_TK-3: ", 0)),
        ("M58", attendance_data.get("Prog_C_Month_9_TK-3: ", 0)),
        ("M59", attendance_data.get("Prog_C_Month_9_4-6: ", 0)),
        ("M60", attendance_data.get("Prog_C_Month_9_7-8: ", 0)),
        ("M61", attendance_data.get("Prog_C_Month_9_9-12: ", 0)),
        ("M57", attendance_data.get("Prog_C_TK_Month_9_TK-3: ", 0)),
        ("N58", attendance_data.get("Prog_C_Month_10_TK-3: ", 0)),
        ("N59", attendance_data.get("Prog_C_Month_10_4-6: ", 0)),
        ("N60", attendance_data.get("Prog_C_Month_10_7-8: ", 0)),
        ("N61", attendance_data.get("Prog_C_Month_10_9-12: ", 0)),
        ("N57", attendance_data.get("Prog_C_TK_Month_10_TK-3: ", 0)),
        ("O58", attendance_data.get("Prog_C_Month_11_TK-3: ", 0)),
        ("O59", attendance_data.get("Prog_C_Month_11_4-6: ", 0)),
        ("O60", attendance_data.get("Prog_C_Month_11_7-8: ", 0)),
        ("O61", attendance_data.get("Prog_C_Month_11_9-12: ", 0)),
        ("O57", attendance_data.get("Prog_C_TK_Month_11_TK-3: ", 0)),
        ("P58", attendance_data.get("Prog_C_Month_12_TK-3: ", 0)),
        ("P59", attendance_data.get("Prog_C_Month_12_4-6: ", 0)),
        ("P60", attendance_data.get("Prog_C_Month_12_7-8: ", 0)),
        ("P61", attendance_data.get("Prog_C_Month_12_9-12: ", 0)),
        ("P57", attendance_data.get("Prog_C_TK_Month_12_TK-3: ", 0)),
        
        # =================================================================
        # PROGRAM N NON-RESIDENT CHARTER PLACEMENTS (Rows 74-78)
        # =================================================================
        ("E75", attendance_data.get("Prog_N_Month_1_TK-3: ", 0)),
        ("E76", attendance_data.get("Prog_N_Month_1_4-6: ", 0)),
        ("E77", attendance_data.get("Prog_N_Month_1_7-8: ", 0)),
        ("E78", attendance_data.get("Prog_N_Month_1_9-12: ", 0)),
        ("E74", attendance_data.get("Prog_N_TK_Month_1_TK-3: ", 0)),
        ("F75", attendance_data.get("Prog_N_Month_2_TK-3: ", 0)),
        ("F76", attendance_data.get("Prog_N_Month_2_4-6: ", 0)),
        ("F77", attendance_data.get("Prog_N_Month_2_7-8: ", 0)),
        ("F78", attendance_data.get("Prog_N_Month_2_9-12: ", 0)),
        ("F74", attendance_data.get("Prog_N_TK_Month_2_TK-3: ", 0)),
        ("G75", attendance_data.get("Prog_N_Month_3_TK-3: ", 0)),
        ("G76", attendance_data.get("Prog_N_Month_3_4-6: ", 0)),
        ("G77", attendance_data.get("Prog_N_Month_3_7-8: ", 0)),
        ("G78", attendance_data.get("Prog_N_Month_3_9-12: ", 0)),
        ("G74", attendance_data.get("Prog_N_TK_Month_3_TK-3: ", 0)),
        ("H75", attendance_data.get("Prog_N_Month_4_TK-3: ", 0)),
        ("H76", attendance_data.get("Prog_N_Month_4_4-6: ", 0)),
        ("H77", attendance_data.get("Prog_N_Month_4_7-8: ", 0)),
        ("H78", attendance_data.get("Prog_N_Month_4_9-12: ", 0)),
        ("H74", attendance_data.get("Prog_N_TK_Month_4_TK-3: ", 0)),
        ("I75", attendance_data.get("Prog_N_Month_5_TK-3: ", 0)),
        ("I76", attendance_data.get("Prog_N_Month_5_4-6: ", 0)),
        ("I77", attendance_data.get("Prog_N_Month_5_7-8: ", 0)),
        ("I78", attendance_data.get("Prog_N_Month_5_9-12: ", 0)),
        ("I74", attendance_data.get("Prog_N_TK_Month_5_TK-3: ", 0)),
        ("J75", attendance_data.get("Prog_N_Month_6_TK-3: ", 0)),
        ("J76", attendance_data.get("Prog_N_Month_6_4-6: ", 0)),
        ("J77", attendance_data.get("Prog_N_Month_6_7-8: ", 0)),
        ("J78", attendance_data.get("Prog_N_Month_6_9-12: ", 0)),
        ("J74", attendance_data.get("Prog_N_TK_Month_6_TK-3: ", 0)),
        ("K75", attendance_data.get("Prog_N_Month_7_TK-3: ", 0)),
        ("K76", attendance_data.get("Prog_N_Month_7_4-6: ", 0)),
        ("K77", attendance_data.get("Prog_N_Month_7_7-8: ", 0)),
        ("K78", attendance_data.get("Prog_N_Month_7_9-12: ", 0)),
        ("K74", attendance_data.get("Prog_N_TK_Month_7_TK-3: ", 0)),
        ("L75", attendance_data.get("Prog_N_Month_8_TK-3: ", 0)),
        ("L76", attendance_data.get("Prog_N_Month_8_4-6: ", 0)),
        ("L77", attendance_data.get("Prog_N_Month_8_7-8: ", 0)),
        ("L78", attendance_data.get("Prog_N_Month_8_9-12: ", 0)),
        ("L74", attendance_data.get("Prog_N_TK_Month_8_TK-3: ", 0)),
        ("M75", attendance_data.get("Prog_N_Month_9_TK-3: ", 0)),
        ("M76", attendance_data.get("Prog_N_Month_9_4-6: ", 0)),
        ("M77", attendance_data.get("Prog_N_Month_9_7-8: ", 0)),
        ("M78", attendance_data.get("Prog_N_Month_9_9-12: ", 0)),
        ("M74", attendance_data.get("Prog_N_TK_Month_9_TK-3: ", 0)),
        ("N75", attendance_data.get("Prog_N_Month_10_TK-3: ", 0)),
        ("N76", attendance_data.get("Prog_N_Month_10_4-6: ", 0)),
        ("N77", attendance_data.get("Prog_N_Month_10_7-8: ", 0)),
        ("N78", attendance_data.get("Prog_N_Month_10_9-12: ", 0)),
        ("N74", attendance_data.get("Prog_N_TK_Month_10_TK-3: ", 0)),
        ("O75", attendance_data.get("Prog_N_Month_11_TK-3: ", 0)),
        ("O76", attendance_data.get("Prog_N_Month_11_4-6: ", 0)),
        ("O77", attendance_data.get("Prog_N_Month_11_7-8: ", 0)),
        ("O78", attendance_data.get("Prog_N_Month_11_9-12: ", 0)),
        ("O74", attendance_data.get("Prog_N_TK_Month_11_TK-3: ", 0)),
        ("P75", attendance_data.get("Prog_N_Month_12_TK-3: ", 0)),
        ("P76", attendance_data.get("Prog_N_Month_12_4-6: ", 0)),
        ("P77", attendance_data.get("Prog_N_Month_12_7-8: ", 0)),
        ("P78", attendance_data.get("Prog_N_Month_12_9-12: ", 0)),
        ("P74", attendance_data.get("Prog_N_TK_Month_12_TK-3: ", 0)),
        
        # =================================================================
        # PROGRAM J INDEPENDENT STUDY CHARTER RESIDENT PLACEMENTS (Rows 64-69)
        # ERROR: CONFLICTS WITH PROGRAM K BELOW
        # ERROR: NEED TO SHIFT PROGRAM J DOWN 1 ROW TO AVOID WRONG DATA
        # =================================================================
        ("E66", attendance_data.get("Prog_J_Month_1_TK-3: ", 0)),
        ("E67", attendance_data.get("Prog_J_Month_1_4-6: ", 0)),
        ("E68", attendance_data.get("Prog_J_Month_1_7-8: ", 0)),
        ("E69", attendance_data.get("Prog_J_Month_1_9-12: ", 0)),
        ("E65", attendance_data.get("Prog_J_TK_Month_1_TK-3: ", 0)),
        ("F66", attendance_data.get("Prog_J_Month_2_TK-3: ", 0)),
        ("F67", attendance_data.get("Prog_J_Month_2_4-6: ", 0)),
        ("F68", attendance_data.get("Prog_J_Month_2_7-8: ", 0)),
        ("F69", attendance_data.get("Prog_J_Month_2_9-12: ", 0)),
        ("F65", attendance_data.get("Prog_J_TK_Month_2_TK-3: ", 0)),
        ("G66", attendance_data.get("Prog_J_Month_3_TK-3: ", 0)),
        ("G67", attendance_data.get("Prog_J_Month_3_4-6: ", 0)),
        ("G68", attendance_data.get("Prog_J_Month_3_7-8: ", 0)),
        ("G69", attendance_data.get("Prog_J_Month_3_9-12: ", 0)),
        ("G65", attendance_data.get("Prog_J_TK_Month_3_TK-3: ", 0)),
        ("H66", attendance_data.get("Prog_J_Month_4_TK-3: ", 0)),
        ("H67", attendance_data.get("Prog_J_Month_4_4-6: ", 0)),
        ("H68", attendance_data.get("Prog_J_Month_4_7-8: ", 0)),
        ("H69", attendance_data.get("Prog_J_Month_4_9-12: ", 0)),
        ("H65", attendance_data.get("Prog_J_TK_Month_4_TK-3: ", 0)),
        ("I66", attendance_data.get("Prog_J_Month_5_TK-3: ", 0)),
        ("I67", attendance_data.get("Prog_J_Month_5_4-6: ", 0)),
        ("I68", attendance_data.get("Prog_J_Month_5_7-8: ", 0)),
        ("I69", attendance_data.get("Prog_J_Month_5_9-12: ", 0)),
        ("I65", attendance_data.get("Prog_J_TK_Month_5_TK-3: ", 0)),
        ("J66", attendance_data.get("Prog_J_Month_6_TK-3: ", 0)),
        ("J67", attendance_data.get("Prog_J_Month_6_4-6: ", 0)),
        ("J68", attendance_data.get("Prog_J_Month_6_7-8: ", 0)),
        ("J69", attendance_data.get("Prog_J_Month_6_9-12: ", 0)),
        ("J65", attendance_data.get("Prog_J_TK_Month_6_TK-3: ", 0)),
        ("K66", attendance_data.get("Prog_J_Month_7_TK-3: ", 0)),
        ("K67", attendance_data.get("Prog_J_Month_7_4-6: ", 0)),
        ("K68", attendance_data.get("Prog_J_Month_7_7-8: ", 0)),
        ("K69", attendance_data.get("Prog_J_Month_7_9-12: ", 0)),
        ("K65", attendance_data.get("Prog_J_TK_Month_7_TK-3: ", 0)),
        ("L66", attendance_data.get("Prog_J_Month_8_TK-3: ", 0)),
        ("L67", attendance_data.get("Prog_J_Month_8_4-6: ", 0)),
        ("L68", attendance_data.get("Prog_J_Month_8_7-8: ", 0)),
        ("L69", attendance_data.get("Prog_J_Month_8_9-12: ", 0)),
        ("L65", attendance_data.get("Prog_J_TK_Month_8_TK-3: ", 0)),
        ("M66", attendance_data.get("Prog_J_Month_9_TK-3: ", 0)),
        ("M67", attendance_data.get("Prog_J_Month_9_4-6: ", 0)),
        ("M68", attendance_data.get("Prog_J_Month_9_7-8: ", 0)),
        ("M69", attendance_data.get("Prog_J_Month_9_9-12: ", 0)),
        ("M65", attendance_data.get("Prog_J_TK_Month_9_TK-3: ", 0)),
        ("N66", attendance_data.get("Prog_J_Month_10_TK-3: ", 0)),
        ("N67", attendance_data.get("Prog_J_Month_10_4-6: ", 0)),
        ("N68", attendance_data.get("Prog_J_Month_10_7-8: ", 0)),
        ("N69", attendance_data.get("Prog_J_Month_10_9-12: ", 0)),
        ("N65", attendance_data.get("Prog_J_TK_Month_10_TK-3: ", 0)),
        ("O66", attendance_data.get("Prog_J_Month_11_TK-3: ", 0)),
        ("O67", attendance_data.get("Prog_J_Month_11_4-6: ", 0)),
        ("O68", attendance_data.get("Prog_J_Month_11_7-8: ", 0)),
        ("O69", attendance_data.get("Prog_J_Month_11_9-12: ", 0)),
        ("O65", attendance_data.get("Prog_J_TK_Month_11_TK-3: ", 0)),
        ("P66", attendance_data.get("Prog_J_Month_12_TK-3: ", 0)),
        ("P67", attendance_data.get("Prog_J_Month_12_4-6: ", 0)),
        ("P68", attendance_data.get("Prog_J_Month_12_7-8: ", 0)),
        ("P69", attendance_data.get("Prog_J_Month_12_9-12: ", 0)),
        ("P65", attendance_data.get("Prog_J_TK_Month_12_TK-3: ", 0)),
        
        # =================================================================
        # PROGRAM K INDEPENDENT STUDY CHARTER NON-RESIDENT PLACEMENTS (82-86)
        # WARNING: These cells OVERWRITE Program N data! Same cells used!
        # =================================================================
        ("E83", attendance_data.get("Prog_K_Month_1_TK-3: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("E84", attendance_data.get("Prog_K_Month_1_4-6: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("E85", attendance_data.get("Prog_K_Month_1_7-8: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("E86", attendance_data.get("Prog_K_Month_1_9-12: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("E82", attendance_data.get("Prog_K_TK_Month_1_TK-3: ", 0)),    # ‚ö†Ô∏è OVERWRITES Prog_N
        ("F83", attendance_data.get("Prog_K_Month_2_TK-3: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("F84", attendance_data.get("Prog_K_Month_2_4-6: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("F85", attendance_data.get("Prog_K_Month_2_7-8: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("F86", attendance_data.get("Prog_K_Month_2_9-12: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("F82", attendance_data.get("Prog_K_TK_Month_2_TK-3: ", 0)),    # ‚ö†Ô∏è OVERWRITES Prog_N
        ("G83", attendance_data.get("Prog_K_Month_3_TK-3: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("G84", attendance_data.get("Prog_K_Month_3_4-6: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("G85", attendance_data.get("Prog_K_Month_3_7-8: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("G86", attendance_data.get("Prog_K_Month_3_9-12: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("G82", attendance_data.get("Prog_K_TK_Month_3_TK-3: ", 0)),    # ‚ö†Ô∏è OVERWRITES Prog_N
        ("H83", attendance_data.get("Prog_K_Month_4_TK-3: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("H84", attendance_data.get("Prog_K_Month_4_4-6: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("H85", attendance_data.get("Prog_K_Month_4_7-8: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("H86", attendance_data.get("Prog_K_Month_4_9-12: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("H82", attendance_data.get("Prog_K_TK_Month_4_TK-3: ", 0)),    # ‚ö†Ô∏è OVERWRITES Prog_N
        ("I83", attendance_data.get("Prog_K_Month_5_TK-3: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("I84", attendance_data.get("Prog_K_Month_5_4-6: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("I85", attendance_data.get("Prog_K_Month_5_7-8: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("I86", attendance_data.get("Prog_K_Month_5_9-12: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("I82", attendance_data.get("Prog_K_TK_Month_5_TK-3: ", 0)),    # ‚ö†Ô∏è OVERWRITES Prog_N
        ("J83", attendance_data.get("Prog_K_Month_6_TK-3: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("J84", attendance_data.get("Prog_K_Month_6_4-6: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("J85", attendance_data.get("Prog_K_Month_6_7-8: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("J86", attendance_data.get("Prog_K_Month_6_9-12: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("J82", attendance_data.get("Prog_K_TK_Month_6_TK-3: ", 0)),    # ‚ö†Ô∏è OVERWRITES Prog_N
        ("K83", attendance_data.get("Prog_K_Month_7_TK-3: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("K84", attendance_data.get("Prog_K_Month_7_4-6: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("K85", attendance_data.get("Prog_K_Month_7_7-8: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("K86", attendance_data.get("Prog_K_Month_7_9-12: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("K82", attendance_data.get("Prog_K_TK_Month_7_TK-3: ", 0)),    # ‚ö†Ô∏è OVERWRITES Prog_N
        ("L83", attendance_data.get("Prog_K_Month_8_TK-3: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("L84", attendance_data.get("Prog_K_Month_8_4-6: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("L85", attendance_data.get("Prog_K_Month_8_7-8: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("L86", attendance_data.get("Prog_K_Month_8_9-12: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("L82", attendance_data.get("Prog_K_TK_Month_8_TK-3: ", 0)),    # ‚ö†Ô∏è OVERWRITES Prog_N
        ("M83", attendance_data.get("Prog_K_Month_9_TK-3: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("M84", attendance_data.get("Prog_K_Month_9_4-6: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("M85", attendance_data.get("Prog_K_Month_9_7-8: ", 0)),        # ‚ö†Ô∏è OVERWRITES Prog_N
        ("M86", attendance_data.get("Prog_K_Month_9_9-12: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("M82", attendance_data.get("Prog_K_TK_Month_9_TK-3: ", 0)),    # ‚ö†Ô∏è OVERWRITES Prog_N
        ("N83", attendance_data.get("Prog_K_Month_10_TK-3: ", 0)),      # ‚ö†Ô∏è OVERWRITES Prog_N
        ("N84", attendance_data.get("Prog_K_Month_10_4-6: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("N85", attendance_data.get("Prog_K_Month_10_7-8: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("N86", attendance_data.get("Prog_K_Month_10_9-12: ", 0)),      # ‚ö†Ô∏è OVERWRITES Prog_N
        ("N82", attendance_data.get("Prog_K_TK_Month_10_TK-3: ", 0)),   # ‚ö†Ô∏è OVERWRITES Prog_N
        ("O83", attendance_data.get("Prog_K_Month_11_TK-3: ", 0)),      # ‚ö†Ô∏è OVERWRITES Prog_N
        ("O84", attendance_data.get("Prog_K_Month_11_4-6: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("O85", attendance_data.get("Prog_K_Month_11_7-8: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("O86", attendance_data.get("Prog_K_Month_11_9-12: ", 0)),      # ‚ö†Ô∏è OVERWRITES Prog_N
        ("O82", attendance_data.get("Prog_K_TK_Month_11_TK-3: ", 0)),   # ‚ö†Ô∏è OVERWRITES Prog_N
        ("P83", attendance_data.get("Prog_K_Month_12_TK-3: ", 0)),      # ‚ö†Ô∏è OVERWRITES Prog_N
        ("P84", attendance_data.get("Prog_K_Month_12_4-6: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("P85", attendance_data.get("Prog_K_Month_12_7-8: ", 0)),       # ‚ö†Ô∏è OVERWRITES Prog_N
        ("P86", attendance_data.get("Prog_K_Month_12_9-12: ", 0)),      # ‚ö†Ô∏è OVERWRITES Prog_N
        ("P82", attendance_data.get("Prog_K_TK_Month_12_TK-3: ", 0)),   # ‚ö†Ô∏è OVERWRITES Prog_N
    ]

    
    # Open the Excel file and prepare to write all data at once
    workbook = openpyxl.load_workbook(output_excel_path)
    worksheet = workbook[output_worksheet_name]
    
    # Write all values to their respective cells in one operation
    for target_cell, attendance_value in cell_mapping_list:
        worksheet[target_cell] = attendance_value
    
    # Save all changes at once
    workbook.save(output_excel_path)
    
    print("‚úÖ All attendance data has been written to Excel successfully!")


# =============================================================================
# MAIN PROGRAM EXECUTION
# =============================================================================

def run_ada_audit_process():
    """
    Main function that orchestrates the entire ADA audit process.
    
    What this does:
    1. Gets user input for Location, School Year, and School Name
    2. Loads student attendance data from Excel
    3. Identifies where each program's data is located
    4. Extracts attendance numbers for each program and month
    5. Consolidates sub-location data with parent programs
    6. Writes the organized data to the audit Excel file
    """
    
    # =================================================================
    # STEP 1: Get user input for Location, School Year, and School Name
    # =================================================================
    print("=" * 60)
    print("üéì ADA AUDIT CONFIGURATION")
    print("=" * 60)
    
    location = input("üìç Enter Location (e.g., TK-8, Elementary, Middle, High): ").strip()
    if not location:
        location = "TK-8"
        print(f"   Using default: {location}")
    
    school_year = input("üìÖ Enter School Year (e.g., 2025-2026, 2024-2025): ").strip()
    if not school_year:
        school_year = "2025-2026"
        print(f"   Using default: {school_year}")
    
    school_name = input("üè´ Enter School Name (e.g., CCCS, Lincoln Elementary): ").strip()
    if not school_name:
        school_name = "CCCS"
        print(f"   Using default: {school_name}")
    
    print(f"\n‚úÖ Configuration:")
    print(f"   Location: {location}")
    print(f"   School Year: {school_year}")
    print(f"   School Name: {school_name}")
    print("=" * 60)
    
    # =================================================================
    # STEP 2: Define file paths and program information
    # =================================================================
    input_attendance_file = (
"C:\\Users\\Shawn\\Downloads\\PrintMonthlyAttendanceSummaryTotals_20251021_143005_82100f5.xlsx"
    )
    output_audit_file = "C:\\Users\\Shawn\\Downloads\\2025-2026_I4C_ADA_Reconciliation.xlsx"
    target_worksheet_name = "Template- Apportionment Summary"
    
    # Define all the program names we're looking for and their short codes
    # Main programs and their sub-locations
    program_name_mappings = {
        # Main Program C locations
        "Program C Charter Resident": "Prog_C",
        "Program C Charter Resident -  Transitional Kindergarten(TK)": "Prog_C_TK",
        "Program C Charter Resident -  McClellan(CM)": "Prog_C_CM",
        "Program C Charter Resident -  Sac Youth Center(SYC)": "Prog_C_SYC",
        
        # Main Program N locations  
        "Program N Non-Resident Charter": "Prog_N", 
        "Program N Non-Resident Charter -  Transitional Kindergarten(TK)": "Prog_N_TK",
        "Program N Non-Resident Charter -  McClellan(CM)": "Prog_N_CM",
        "Program N Non-Resident Charter -  Sac Youth Center(SYC)": "Prog_N_SYC",
        
        # Independent Study programs
        "Program J Indep Study Charter Resident": "Prog_J",
        "Program J Indep Study Charter Non-Resident -  Transitional Kindergarten(TK)": "Prog_J_TK",
        "Program K Indep Study Charter Non-Resident": "Prog_K",
        "Program K Indep Study Charter Non-Resident -  Transitional Kindergarten(TK)": "Prog_K_TK",
    }
    
    # Define which sub-programs should be combined with their parent programs
    program_consolidation_rules = {
        "Prog_C": ["Prog_C", "Prog_C_CM", "Prog_C_SYC"],  # Combine main C + CM + SYC
        "Prog_C_TK": ["Prog_C_TK"],  # TK stays separate
        "Prog_N": ["Prog_N", "Prog_N_CM", "Prog_N_SYC"],  # Combine main N + CM + SYC  
        "Prog_N_TK": ["Prog_N_TK"],  # TK stays separate
        "Prog_J": ["Prog_J"],
        "Prog_J_TK": ["Prog_J_TK"],
        "Prog_K": ["Prog_K"],
        "Prog_K_TK": ["Prog_K_TK"],
    }
    
    # =================================================================
    # STEP 3: Load the attendance data from Excel
    # =================================================================
    print("üìä Loading student attendance data from Excel...")
    student_attendance_data = pd.read_excel(input_attendance_file, header=None)
    
    # =================================================================
    # STEP 4: Find where each program's data starts and ends
    # =================================================================
    print("üîç Locating program boundaries in the data...")
    
    program_boundaries = {}
    for short_code in program_name_mappings.values():
        program_boundaries[short_code] = {"start": None, "stop": None}
    
    # Find the row ranges for each program
    for full_program_name, short_code in program_name_mappings.items():
        matching_rows = find_rows_containing_program_name(student_attendance_data, full_program_name)
        start_row, end_row = find_program_boundary_rows(matching_rows)
        program_boundaries[short_code]["start"] = start_row
        program_boundaries[short_code]["stop"] = end_row
    
    # =================================================================
    # STEP 5: Adjust boundaries to prevent overlaps
    # =================================================================
    print("üîß Adjusting program boundaries to prevent overlaps...")

    # Fix Program C boundaries
    prog_C_tk_start = program_boundaries["Prog_C_TK"]["start"]
    prog_N_start = program_boundaries["Prog_N"]["start"]

    if prog_C_tk_start is not None and prog_N_start is not None:
        program_boundaries["Prog_C"]["stop"] = prog_C_tk_start - 1

    if prog_N_start is not None:
        program_boundaries["Prog_C_TK"]["stop"] = prog_N_start - 1

    # Fix Program N boundaries
    prog_N_tk_start = program_boundaries["Prog_N_TK"]["start"]
    if prog_N_tk_start is not None:
        program_boundaries["Prog_N"]["stop"] = prog_N_tk_start - 1

    # Fix remaining program boundaries
    programs_to_adjust = ["Prog_N_TK", "Prog_J", "Prog_K"]
    for i in range(len(programs_to_adjust) - 1):
        current_program = programs_to_adjust[i]
        next_program = programs_to_adjust[i + 1]
        
        current_start = program_boundaries[current_program]["start"]
        next_start = program_boundaries[next_program]["start"]
        
        if current_start is not None and next_start is not None:
            program_boundaries[current_program]["stop"] = next_start - 1
    
    # =================================================================
    # STEP 6: Display boundaries and allow user verification
    # =================================================================
    print("\nüìç Program boundary information:")
    for program_code, boundaries in program_boundaries.items():
        start = boundaries.get("start", "Not found")
        stop = boundaries.get("stop", "Not found") 
        print(f"  {program_code}: Start Row {start}, End Row {stop}")
    
    # Allow user to verify and correct boundaries if needed
    for program_code in program_boundaries.keys():
        user_response = input(
            f"\n‚ùì Are the boundaries for {program_code} correct? (yes/no): "
        ).lower().strip()
        
        if user_response == "no":
            while True:
                user_input = input(
                    f"üìù Enter new start and end rows for {program_code} (format: start,end): "
                )
                try:
                    boundary_parts = user_input.split(",")
                    if len(boundary_parts) == 2:
                        new_start = (
                            int(boundary_parts[0].strip()) 
                            if boundary_parts[0].strip().lower() != "none" 
                            else None
                        )
                        new_end = (
                            int(boundary_parts[1].strip())
                            if boundary_parts[1].strip().lower() != "none"
                            else None
                        )
                        program_boundaries[program_code]["start"] = new_start
                        program_boundaries[program_code]["stop"] = new_end
                        print(f"‚úÖ Updated {program_code} boundaries: Start {new_start}, End {new_end}")
                        break
                    else:
                        print("‚ùå Please enter exactly two values separated by a comma.")
                except ValueError:
                    print("‚ùå Please enter valid numbers or 'none'.")
    
    # =================================================================
    # STEP 7: Find all month occurrences in the data
    # =================================================================
    print("\nüìÖ Finding month occurrences in attendance data...")
    
    monthly_attendance_by_program = {}
    for month_number in range(1, 13):  # Months 1-12
        rows_with_this_month = find_rows_containing_month_number(student_attendance_data, month_number)
        monthly_attendance_by_program[month_number] = rows_with_this_month
        print(f"  Month {month_number}: Found in {len(rows_with_this_month)} rows")
    
    # =================================================================
    # STEP 8: Extract all raw attendance data (including sub-locations)
    # =================================================================
    print("\nüìà Extracting attendance data for all programs and months...")
    
    raw_attendance_data = extract_student_attendance_data(
        monthly_attendance_by_program, 
        program_boundaries, 
        student_attendance_data
    )
    
    print(f"‚úÖ Extracted {len(raw_attendance_data)} raw attendance data points")
    
    # =================================================================
    # STEP 9: Consolidate sub-location data with parent programs
    # =================================================================
    print("\nüîÑ Consolidating sub-location data with parent programs...")
    print("   Program C Total = Main Program C + McClellan (CM) + Sac Youth Center (SYC)")
    print("   Program N Total = Main Program N + McClellan (CM) + Sac Youth Center (SYC)")
    
    consolidated_attendance_data = {}
    
    # Process each consolidation rule
    for parent_program, child_programs in program_consolidation_rules.items():
        print(f"  Consolidating {parent_program}: {child_programs}")
        
        # For each month (1-12) and age group combination
        for month in range(1, 13):
            for age_group in ["TK-3", "4-6", "7-8", "9-12"]:
                # Create the field name pattern
                field_pattern = f"{parent_program}_Month_{month}_{age_group}: "
                
                # Sum up values from all child programs
                total_value = 0
                found_values = []
                
                for child_program in child_programs:
                    child_field_pattern = f"{child_program}_Month_{month}_{age_group}: "
                    child_value = raw_attendance_data.get(child_field_pattern, 0)
                    
                    if child_value and not pd.isna(child_value) and child_value != 0:
                        total_value += child_value
                        found_values.append(f"{child_program}: {child_value}")
                
                # Store the consolidated value
                consolidated_attendance_data[field_pattern] = total_value
                
                # Log consolidation details for non-zero values
                if total_value > 0:
                    print(f"    {field_pattern} = {' + '.join(found_values)} = {total_value}")
    
    print(f"‚úÖ Consolidated {len(consolidated_attendance_data)} attendance data points")
    
    # =================================================================
    # DEBUG: Show all keys in consolidated data
    # =================================================================
    print("\nüîç DEBUG: Checking all keys in consolidated_attendance_data...")
    print(f"Total keys: {len(consolidated_attendance_data)}")
    
    # Group by program to see what we have
    program_keys = {}
    for key in consolidated_attendance_data.keys():
        program = key.split('_Month_')[0] if '_Month_' in key else 'Unknown'
        if program not in program_keys:
            program_keys[program] = []
        program_keys[program].append(key)
    
    print("\nKeys grouped by program:")
    for program, keys in sorted(program_keys.items()):
        print(f"  {program}: {len(keys)} keys")
        # Show first 3 examples
        for key in keys[:3]:
            value = consolidated_attendance_data[key]
            print(f"    - {key} = {value}")
        if len(keys) > 3:
            print(f"    ... and {len(keys) - 3} more")
    
    # Check for keys NOT in cell_mapping_list
    print("\n‚ö†Ô∏è  Checking for keys that are NOT in cell_mapping_list...")
    expected_keys = set([
        # Prog_C
        *[f"Prog_C_Month_{m}_TK-3: " for m in range(1, 13)],
        *[f"Prog_C_Month_{m}_4-6: " for m in range(1, 13)],
        *[f"Prog_C_Month_{m}_7-8: " for m in range(1, 13)],
        *[f"Prog_C_Month_{m}_9-12: " for m in range(1, 13)],
        *[f"Prog_C_TK_Month_{m}_TK-3: " for m in range(1, 13)],
        # Prog_N
        *[f"Prog_N_Month_{m}_TK-3: " for m in range(1, 13)],
        *[f"Prog_N_Month_{m}_4-6: " for m in range(1, 13)],
        *[f"Prog_N_Month_{m}_7-8: " for m in range(1, 13)],
        *[f"Prog_N_Month_{m}_9-12: " for m in range(1, 13)],
        *[f"Prog_N_TK_Month_{m}_TK-3: " for m in range(1, 13)],
        # Prog_J
        *[f"Prog_J_Month_{m}_TK-3: " for m in range(1, 13)],
        *[f"Prog_J_Month_{m}_4-6: " for m in range(1, 13)],
        *[f"Prog_J_Month_{m}_7-8: " for m in range(1, 13)],
        *[f"Prog_J_Month_{m}_9-12: " for m in range(1, 13)],
        *[f"Prog_J_TK_Month_{m}_TK-3: " for m in range(1, 13)],
        # Prog_K
        *[f"Prog_K_Month_{m}_TK-3: " for m in range(1, 13)],
        *[f"Prog_K_Month_{m}_4-6: " for m in range(1, 13)],
        *[f"Prog_K_Month_{m}_7-8: " for m in range(1, 13)],
        *[f"Prog_K_Month_{m}_9-12: " for m in range(1, 13)],
        *[f"Prog_K_TK_Month_{m}_TK-3: " for m in range(1, 13)],
    ])
    
    unexpected_keys = []
    for key in consolidated_attendance_data.keys():
        if key not in expected_keys:
            unexpected_keys.append(key)
    
    if unexpected_keys:
        print(f"Found {len(unexpected_keys)} unexpected keys:")
        for key in unexpected_keys[:20]:  # Show first 20
            value = consolidated_attendance_data[key]
            print(f"  ‚ùå {key} = {value}")
        if len(unexpected_keys) > 20:
            print(f"  ... and {len(unexpected_keys) - 20} more")
    else:
        print("‚úÖ All keys are expected and mapped!")
    
    # =================================================================
    # STEP 10: Write consolidated data to Excel audit file
    # =================================================================
    print("\nüíæ Writing consolidated attendance data to Excel audit file...")
    
    write_all_attendance_data_to_excel_efficiently(
        consolidated_attendance_data, 
        output_audit_file, 
        target_worksheet_name
    )
    
    print("\nüéâ ADA Audit process completed successfully!")
    print(f"üìä Results saved to: {output_audit_file}")
    print(f"üìç Configuration used: {location}, {school_year}, {school_name}")
    print("\nüí° Note: McClellan (CM) and Sac Youth Center (SYC) totals have been")
    print("   automatically added to their respective parent program totals.")


# =============================================================================
# RUN THE PROGRAM
# =============================================================================

if __name__ == "__main__":
    run_ada_audit_process()
